# -*- coding: utf-8 -*-
"""
该目录下存放一些公关文件  例如读写Excel的库
"""
from openpyxl import load_workbook
import os
import shutil


class Reader:

    def __init__(self):
        self.wb = None
        self.ws = None
        self.rows = 0
        self.column = 0
        self.lines = []

    # 打开Excel文件
    def open_excel(self, filename):
        # 如果打开的文件不存在  给出提示
        if not os.path.isfile(filename):
            print('error: %s not exists!' % filename)

        # 获取需要读取的工作簿对象
        self.wb = load_workbook(filename)
        self.set_sheet(self.wb.sheetnames[0])

    # 获取需要读取的sheet表单
    def get_sheets(self):
        return self.wb.sheetnames

    # 设置当前表单的数据
    def set_sheet(self, sheet_name):
        self.ws = self.wb[sheet_name]
        self.__read_line()

    # 读取每行的数据
    def __read_line(self):
        # 每次添加数据前把上次的数据清空  防止重复
        self.lines = []
        for row in self.ws.rows:
            temp = []
            for cell in row:
                if cell.value is None:
                    cell.value = ''
                temp.append(cell.value)
            self.lines.append(temp)


# 向Excel文件写入的库
class Writer:
    def __init__(self):
        self.wb = None
        self.ws = None
        self.rows = 0
        self.destFile = None

    def copy_open(self, source_name, copy_name):
        # 判断要复制的文件是否存在
        if not os.path.isfile(source_name):
            print(source_name+'not exists!')
            return

        # 判断要新建的文档是否存在，存在则提示
        if os.path.isfile(copy_name):
            print('warning: ' + copy_name + 'is already exist!')

        # 从源文件复制数据到新文件  并记录新文件
        self.destFile = shutil.copyfile(os.path.join(os.path.abspath('.'), source_name),
                                        os.path.join(os.path.abspath('.'), copy_name))

        # 获取工作簿对象
        self.wb = load_workbook(self.destFile)
        # 默认的工作表对象
        self.set_active()

    # 切换sheet页面
    def set_sheet(self, sheet_name):
        self.ws = self.wb[sheet_name]

    def set_active(self):
        """
        把工作簿定位到第一个sheet
        """
        self.ws = self.wb[self.wb.sheetnames[0]]

    # 写入指定的单元格  保留原格式(格式问题先忽略)
    def write(self, r, c, value):
        # 获取要写入的单元格
        try:
            self.ws.cell(row=r, column=c).value = value
        except:
            self.ws.cell(row=r, column=c).value = 'WriteFail'

    # 保存文件
    def save_close(self):
        self.wb.save(self.destFile)
